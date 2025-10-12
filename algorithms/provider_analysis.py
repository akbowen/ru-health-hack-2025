import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook

class ProviderAnalyzer:
    """Analyze provider utilization from generated schedules"""
    
    def __init__(self, output_dir="output"):
        self.output_dir = output_dir
        self.schedule_files = {
            'Minimize Providers': 'schedule_1_minimize_providers.xlsx',
            'Balanced Providers': 'schedule_2_balanced_providers.xlsx',
            'Phase 1 Conservative': 'schedule_3_phase1_conservative.xlsx'
        }
    
    def analyze_schedule(self, filepath, variation_name):
        """Analyze a single schedule file and extract provider statistics"""
        
        print(f"\nAnalyzing: {variation_name}")
        print(f"File: {filepath}")
        
        # Load workbook
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Find header row (skip satisfaction score rows)
        header_row = None
        for row_idx in range(1, 20):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value == "Day":
                header_row = row_idx
                break
        
        if header_row is None:
            print("  ⚠️  Could not find header row")
            return None
        
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(header_row, col).value
            if header:
                headers.append(header)
        
        # Parse site-shift columns
        site_shift_cols = []
        for col_idx, header in enumerate(headers[1:], start=2):  # Skip "Day" column
            if " - " in str(header):
                parts = header.split(" - ")
                if len(parts) == 2:
                    site, shift = parts
                    site_shift_cols.append((col_idx, site.strip(), shift.strip()))
        
        print(f"  Found {len(site_shift_cols)} site-shift columns")
        
        # Count provider assignments
        provider_stats = {}  # provider -> {total, MD1, MD2, PM, shifts_list}
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            day = ws.cell(row_idx, 1).value
            if not day or not isinstance(day, int):
                continue
            
            for col_idx, site, shift in site_shift_cols:
                cell_value = ws.cell(row_idx, col_idx).value
                
                if not cell_value or cell_value == "UNCOVERED":
                    continue
                
                # Handle dual providers (e.g., "John & Mary")
                if " & " in str(cell_value):
                    providers_in_cell = [p.strip() for p in str(cell_value).split(" & ")]
                else:
                    providers_in_cell = [str(cell_value).strip()]
                
                for provider in providers_in_cell:
                    if provider not in provider_stats:
                        provider_stats[provider] = {
                            'total': 0,
                            'MD1': 0,
                            'MD2': 0,
                            'PM': 0,
                            'shifts': set(),  # Unique (shift, day) combinations
                            'sites': 0
                        }
                    
                    # Track unique shift-day (multi-site on same shift = 1 shift)
                    shift_day = (shift, day)
                    if shift_day not in provider_stats[provider]['shifts']:
                        provider_stats[provider]['shifts'].add(shift_day)
                        provider_stats[provider]['total'] += 1
                        provider_stats[provider][shift] += 1
                    
                    # Count sites
                    provider_stats[provider]['sites'] += 1
        
        print(f"  Providers found: {len(provider_stats)}")
        
        # Sort by total shifts (descending)
        sorted_providers = sorted(provider_stats.items(), 
                                 key=lambda x: x[1]['total'], 
                                 reverse=True)
        
        # Get top 5
        top5 = sorted_providers[:5]
        
        print(f"\n  Top 5 Providers by Shift Count:")
        for rank, (provider, stats) in enumerate(top5, 1):
            print(f"    {rank}. {provider}")
            print(f"       Total Shifts: {stats['total']} "
                  f"(MD1: {stats['MD1']}, MD2: {stats['MD2']}, PM: {stats['PM']})")
            print(f"       Total Sites: {stats['sites']} "
                  f"(Avg: {stats['sites']/stats['total']:.1f} sites/shift)")
        
        return {
            'variation': variation_name,
            'total_providers': len(provider_stats),
            'top5': top5,
            'all_stats': provider_stats
        }
    
    def create_top5_report(self, analyses):
        """Create Excel report with top 5 providers for each variation"""
        
        print("\n" + "=" * 80)
        print("CREATING TOP 5 PROVIDER REPORT")
        print("=" * 80)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Top 5 Analysis"
        
        # Title
        ws.append(["TOP 5 PROVIDERS BY SHIFT COUNT - COMPARISON"])
        ws.cell(1, 1).font = Font(bold=True, size=16, color="0000FF")
        ws.merge_cells('A1:H1')
        ws.append([])
        
        row_num = 3
        
        for analysis in analyses:
            if analysis is None:
                continue
            
            variation = analysis['variation']
            top5 = analysis['top5']
            total_providers = analysis['total_providers']
            
            # Variation header
            ws.append([f"{variation} (Total Providers: {total_providers})"])
            ws.cell(row_num, 1).font = Font(bold=True, size=14, color="FF0000")
            ws.merge_cells(f'A{row_num}:H{row_num}')
            row_num += 1
            
            # Column headers
            headers = ["Rank", "Provider Name", "Total Shifts", "MD1", "MD2", "PM", 
                      "Total Sites", "Sites/Shift"]
            ws.append(headers)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row_num, col_idx)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            row_num += 1
            
            # Top 5 data
            for rank, (provider, stats) in enumerate(top5, 1):
                avg_sites = stats['sites'] / stats['total'] if stats['total'] > 0 else 0
                
                row_data = [
                    rank,
                    provider,
                    stats['total'],
                    stats['MD1'],
                    stats['MD2'],
                    stats['PM'],
                    stats['sites'],
                    f"{avg_sites:.1f}"
                ]
                
                ws.append(row_data)
                
                # Highlight rank 1
                if rank == 1:
                    for col in range(1, 9):
                        ws.cell(row_num, col).fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid"
                        )
                
                row_num += 1
            
            # Blank rows between variations
            ws.append([])
            ws.append([])
            row_num += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Save
        report_path = f"{self.output_dir}/top5_provider_analysis.xlsx"
        wb.save(report_path)
        
        print(f"\n✓ Report created: {report_path}")
        return report_path
    
    def run_analysis(self):
        """Run complete analysis"""
        
        print("\n" + "=" * 80)
        print("PROVIDER UTILIZATION ANALYSIS")
        print("=" * 80)
        print("Analyzing top 5 providers by shift count for each variation\n")
        
        analyses = []
        
        for variation_name, filename in self.schedule_files.items():
            filepath = f"{self.output_dir}/{filename}"
            
            try:
                analysis = self.analyze_schedule(filepath, variation_name)
                analyses.append(analysis)
            except Exception as e:
                print(f"  ⚠️  Error analyzing {variation_name}: {e}")
                analyses.append(None)
        
        # Create comparison report
        report_path = self.create_top5_report(analyses)
        
        # Print summary comparison
        print("\n" + "=" * 80)
        print("SUMMARY COMPARISON")
        print("=" * 80)
        
        for analysis in analyses:
            if analysis is None:
                continue
            
            variation = analysis['variation']
            top5 = analysis['top5']
            total_prov = analysis['total_providers']
            
            if top5:
                top1_name, top1_stats = top5[0]
                print(f"\n{variation}:")
                print(f"  Total providers utilized: {total_prov}")
                print(f"  #1 busiest: {top1_name} - {top1_stats['total']} shifts")
                print(f"  Top 5 average: {sum(s[1]['total'] for s in top5) / 5:.1f} shifts")
                
                # Show distribution
                top5_total = sum(s[1]['total'] for s in top5)
                all_total = sum(s['total'] for s in analysis['all_stats'].values())
                concentration = top5_total / all_total * 100 if all_total > 0 else 0
                print(f"  Top 5 concentration: {concentration:.1f}% of all shifts")


# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    analyzer = ProviderAnalyzer(output_dir="output")
    analyzer.run_analysis()

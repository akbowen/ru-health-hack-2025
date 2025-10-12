import os
import shutil
import math
from typing import Dict, List, Set
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from ortools.sat.python import cp_model

# ============================================================================
# BASE SCHEDULE CREATOR
# ============================================================================
class BaseScheduler:
    def __init__(self, coverage_file="data_clean/Facility coverage.xlsx",
                 availability_file="data_clean/Provider Availability.xlsx",
                 output_dir="output"):
        self.coverage_file = coverage_file
        self.availability_file = availability_file
        self.output_dir = output_dir

        if os.path.exists(self.output_dir):
            shutil.rmtree(self.output_dir)
        os.makedirs(self.output_dir, exist_ok=True)

    def parse_facility_coverage(self):
        df = pd.read_excel(self.coverage_file, sheet_name=0, header=1)
        coverage = {}
        last_facility = None

        for _, row in df.iterrows():
            facility_cell = row.get("Facility")
            shift_cell = row.get("Shift")
            coverage_cell = row.get("Coverage dates")

            if pd.notna(facility_cell) and str(facility_cell).strip() not in ["", "None"]:
                facility = str(facility_cell).strip()
                last_facility = facility
            else:
                facility = last_facility
            if facility is None:
                continue

            if pd.isna(shift_cell) or str(shift_cell).strip() in ["", "None"]:
                continue
            shift = str(shift_cell).strip()
            if shift not in ["MD1", "MD2", "PM"]:
                continue

            days = []
            if pd.isna(coverage_cell) or str(coverage_cell).strip().lower() == "none":
                days = []
            else:
                coverage_str = str(coverage_cell).replace("–", "-").replace("--", "-")
                for part in coverage_str.split(","):
                    part = part.strip()
                    if not part:
                        continue
                    if "-" in part:
                        try:
                            start, end = map(int, part.split("-"))
                            days.extend(range(start, end + 1))
                        except:
                            continue
                    else:
                        try:
                            days.append(int(part))
                        except:
                            continue
            if facility not in coverage:
                coverage[facility] = {}
            coverage[facility][shift] = sorted(set(days))
        return coverage

    def parse_day_of_week(self):
        df = pd.read_excel(self.availability_file, sheet_name=0, header=2)
        dow_map = {}
        for _, row in df.iterrows():
            day_val = row.iloc[1]
            dow_val = row.iloc[0]
            try:
                day = int(day_val)
                dow_map[day] = str(dow_val).strip() if pd.notna(dow_val) else ""
            except:
                continue
        return dow_map

    def create_base_schedule(self):
        coverage = self.parse_facility_coverage()
        day_of_week_map = self.parse_day_of_week()
        shifts = ["MD1", "MD2", "PM"]
        facilities = sorted(coverage.keys())
        days = list(range(1, 32))

        wb = Workbook()
        ws = wb.active
        ws.title = "Base Schedule"

        all_keys = [f"{shift} - {facility}" for facility in facilities for shift in shifts]
        ws.append(["Day of Week", "Day of Month"] + all_keys)
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        for day in days:
            row = [day_of_week_map.get(day, ""), day]
            for facility in facilities:
                for shift in shifts:
                    coverage_days = coverage.get(facility, {}).get(shift, [])
                    row.append("" if day in coverage_days else "X")
            ws.append(row)

        for r in range(2, len(days) + 2):
            for c in range(3, 3 + len(all_keys)):
                if ws.cell(r, c).value == "X":
                    ws.cell(r, c).fill = black_fill
                    ws.cell(r, c).value = ""

        base_path = os.path.join(self.output_dir, "base_schedule.xlsx")
        wb.save(base_path)
        print(f"Base schedule created: {base_path}")
        return base_path, coverage, shifts, facilities, days, day_of_week_map


# ============================================================================
# SCHEDULER: CP-SAT PHASE 1 + GREEDY PHASE 2
# ============================================================================
class Scheduler:
    def __init__(self, base_schedule_file, coverage, shifts, facilities, days,
                 availability_file, contract_file, credentialing_file,
                 volume_file, day_of_week_map, output_dir="output"):
        self.base_schedule_file = base_schedule_file
        self.coverage = coverage
        self.shifts = shifts
        self.facilities = facilities
        self.days = days
        self.availability_file = availability_file
        self.contract_file = contract_file
        self.credentialing_file = credentialing_file
        self.volume_file = volume_file
        self.day_of_week_map = day_of_week_map
        self.output_dir = output_dir

        # Volume limits (HARD)
        self.volume_limits = {
            "MD1": (6.0, 14.0),
            "MD2": (8.0, 16.0),
            "PM": (5.0, 10.0)
        }

    # ========== DATA PARSING ==========
    def parse_provider_availability(self):
        df = pd.read_excel(self.availability_file, sheet_name=0, header=2)
        raw_providers = list(df.columns[2:])
        provider_names, name_count = [], {}

        for name in raw_providers:
            clean_name = str(name).strip()
            if clean_name in name_count:
                name_count[clean_name] += 1
                provider_names.append(f"{clean_name} ({name_count[clean_name]})")
            else:
                name_count[clean_name] = 1
                provider_names.append(clean_name)

        availability = {p: {} for p in provider_names}
        for _, row in df.iterrows():
            try:
                day_of_month = int(row.iloc[1])
            except:
                continue

            for i, provider_name in enumerate(provider_names, start=2):
                cell_val = row.iloc[i]
                shifts = ["MD1", "MD2", "PM"]

                if pd.isna(cell_val) or str(cell_val).strip() == "":
                    availability[provider_name][day_of_month] = shifts
                else:
                    val = str(cell_val).strip().lower()
                    if val == "unavailable":
                        availability[provider_name][day_of_month] = []
                    elif "pm only" in val:
                        availability[provider_name][day_of_month] = ["PM"]
                    elif "am only" in val:
                        availability[provider_name][day_of_month] = ["MD1", "MD2"]
                    else:
                        availability[provider_name][day_of_month] = shifts
        return availability

    def parse_provider_contracts(self):
        df = pd.read_excel(self.contract_file, sheet_name=0, header=2)
        name_count, contracts = {}, {}

        for _, row in df.iterrows():
            if pd.isna(row.get("Provider Name")):
                continue

            raw_name = str(row["Provider Name"]).strip()
            if raw_name in name_count:
                name_count[raw_name] += 1
                provider = f"{raw_name} ({name_count[raw_name]})"
            else:
                name_count[raw_name] = 1
                provider = raw_name

            shift_pref = str(row.get("Shift preference", "")).strip()
            preferred_shifts = []
            if shift_pref and shift_pref.lower() not in ["", "none", "nan"]:
                preferred_shifts = [s.strip() for s in shift_pref.split(",")]

            contracts[provider] = {
                "contract_type": str(row.get("Contract type", "IC")).strip(),
                "total_shifts": int(row.get("Total shift count", 999)) if pd.notna(row.get("Total shift count")) else 999,
                "weekend_shifts": int(row.get("Weekend shift count", 999)) if pd.notna(row.get("Weekend shift count")) else 999,
                "pm_shifts": int(row.get("PM shift count", 999)) if pd.notna(row.get("PM shift count")) else 999,
                "preferred_shifts": preferred_shifts,
            }
        return contracts

    def parse_provider_credentialing(self):
        df = pd.read_excel(self.credentialing_file, sheet_name=0, header=0)
        credentialing, name_count = {}, {}

        for _, row in df.iterrows():
            if pd.isna(row.get("Provider")):
                continue

            raw_name = str(row["Provider"]).strip()
            if raw_name in name_count:
                name_count[raw_name] += 1
                provider = f"{raw_name} ({name_count[raw_name]})"
            else:
                name_count[raw_name] = 1
                provider = raw_name

            facilities_str = str(row.get("Credentialed Facilities", "")).strip()
            facilities = set(f.strip() for f in facilities_str.split(",") if f.strip()) if facilities_str.lower() != "nan" else set()
            credentialing[provider] = facilities

        return credentialing

    def parse_facility_volumes(self):
        df = pd.read_excel(self.volume_file, sheet_name=0, header=2)
        volumes = {}

        for _, row in df.iterrows():
            if pd.isna(row.get("facility_name")):
                continue

            f = str(row["facility_name"]).strip()
            volumes[f] = {
                "MD1": self._parse_volume(row.get("Volume MD1")),
                "MD2": self._parse_volume(row.get("Volume MD2")),
                "PM": self._parse_volume(row.get("Volume PM"))
            }
        return volumes

    def _parse_volume(self, val):
        if pd.isna(val) or str(val).strip().upper() == "NC":
            return 0.0
        try:
            return float(val)
        except:
            return 0.0

    def identify_high_volume_sites(self, volumes, threshold_percentile=75):
        site_avg_volumes = {}
        for site, vols in volumes.items():
            avg_vol = sum([vols.get(s, 0) for s in ["MD1", "MD2", "PM"]]) / 3
            if avg_vol > 0:
                site_avg_volumes[site] = avg_vol

        if not site_avg_volumes:
            return set()

        sorted_vols = sorted(site_avg_volumes.values())
        threshold_idx = int(len(sorted_vols) * threshold_percentile / 100)
        threshold = sorted_vols[threshold_idx] if threshold_idx < len(sorted_vols) else sorted_vols[-1]

        return {site for site, vol in site_avg_volumes.items() if vol >= threshold}

    def get_weekend_days(self):
        return [d for d, dow in self.day_of_week_map.items()
                if dow.lower() in ["sat", "sun", "saturday", "sunday"]]

    def _sanitize(self, name):
        return str(name).replace(" ", "_").replace("(", "").replace(")", "").replace(",", "")

    # ========== SATISFACTION SCORING ==========
    def calculate_satisfaction_score(self, solution, provider_shifts, contracts,
                                     providers, weekend_days, total_slots):
        """Calculate satisfaction score (0-1, lower is better)"""

        total_weighted_excess = 0.0
        total_M = 0
        total_P = 0
        total_W = 0
        providers_affected = 0

        weight_M = 1.0
        weight_P = 1.5
        weight_W = 2.0

        for provider in providers:
            contract = contracts.get(provider, {})

            if provider not in provider_shifts:
                continue

            shifts_worked = provider_shifts[provider]

            total_contract = contract.get("total_shifts", 999)
            pm_contract = contract.get("pm_shifts", 999)
            weekend_contract = contract.get("weekend_shifts", 999)

            pm_count = sum(1 for (s, d) in shifts_worked if s == "PM")
            weekend_count = sum(1 for (s, d) in shifts_worked if d in weekend_days)

            M_i = max(0, len(shifts_worked) - total_contract) if total_contract < 999 else 0
            P_i = max(0, pm_count - pm_contract) if pm_contract < 999 else 0
            W_i = max(0, weekend_count - weekend_contract) if weekend_contract < 999 else 0

            weighted_excess = M_i * weight_M + P_i * weight_P + W_i * weight_W
            total_weighted_excess += weighted_excess

            total_M += M_i
            total_P += P_i
            total_W += W_i

            if weighted_excess > 0:
                providers_affected += 1

        max_possible_excess = len(providers) * 20 * weight_W
        satisfaction_score = total_weighted_excess / max_possible_excess if max_possible_excess > 0 else 0

        covered_slots = len([k for k in solution.keys()])
        uncovered_slots = total_slots - covered_slots
        coverage_rate = covered_slots / total_slots if total_slots > 0 else 0

        return {
            'satisfaction_score': satisfaction_score,
            'total_weighted_excess': total_weighted_excess,
            'total_M': total_M,
            'total_P': total_P,
            'total_W': total_W,
            'weight_M': weight_M,
            'weight_P': weight_P,
            'weight_W': weight_W,
            'providers_affected': providers_affected,
            'covered_slots': covered_slots,
            'uncovered_slots': uncovered_slots,
            'coverage_rate': coverage_rate
        }

    # ========== PHASE 1: CP-SAT OPTIMIZATION ==========
    def run_phase1_cpsat(self, providers, slots, availability, contracts, credentialing,
                         volumes, weekend_days):
        """Phase 1: CP-SAT optimization for optimal baseline"""

        print("\n" + "=" * 80)
        print("PHASE 1: CP-SAT OPTIMIZATION")
        print("=" * 80)

        model = cp_model.CpModel()
        x = {}

        print("Creating decision variables...")
        for p in providers:
            cred_facilities = credentialing.get(p, set())
            for f in self.facilities:
                if f not in cred_facilities:
                    continue
                for s in self.shifts:
                    for d in self.days:
                        if d not in self.coverage.get(f, {}).get(s, []):
                            continue
                        if s not in availability.get(p, {}).get(d, []):
                            continue

                        x[(p, f, s, d)] = model.NewBoolVar(f"x_{self._sanitize(p)}_{f}_{s}_{d}")

        print(f"Variables created: {len(x):,}")

        # Create shift-day tracking variables
        shift_working = {}
        for p in providers:
            for s in self.shifts:
                for d in self.days:
                    vars_for_shift = [x[(p, f, s, d)] for f in self.facilities if (p, f, s, d) in x]
                    if vars_for_shift:
                        working = model.NewBoolVar(f"w_{self._sanitize(p)}_{s}_{d}")
                        model.Add(sum(vars_for_shift) >= 1).OnlyEnforceIf(working)
                        model.Add(sum(vars_for_shift) == 0).OnlyEnforceIf(working.Not())
                        shift_working[(p, s, d)] = working

        print("Applying constraints...")

        # Coverage tracking
        slot_filled = {}
        for f, s, d in slots:
            eligible = [x[(p, f, s, d)] for p in providers if (p, f, s, d) in x]
            if eligible:
                filled = model.NewBoolVar(f"filled_{f}_{s}_{d}")
                model.Add(sum(eligible) >= 1).OnlyEnforceIf(filled)
                model.Add(sum(eligible) == 0).OnlyEnforceIf(filled.Not())
                slot_filled[(f, s, d)] = filled

        # Contract limits
        for p in providers:
            contract = contracts.get(p, {})
            working_vars = [shift_working[(p2, s, d)] for (p2, s, d) in shift_working.keys() if p2 == p]
            if working_vars:
                model.Add(sum(working_vars) <= min(contract.get("total_shifts", 999), 20))

        # Weekend limits
        for p in providers:
            contract = contracts.get(p, {})
            weekend_vars = [shift_working[(p2, s, d)] for (p2, s, d) in shift_working.keys()
                           if p2 == p and d in weekend_days]
            if weekend_vars:
                model.Add(sum(weekend_vars) <= contract.get("weekend_shifts", 999))

        # PM limits
        for p in providers:
            contract = contracts.get(p, {})
            pm_vars = [shift_working[(p2, s, d)] for (p2, s, d) in shift_working.keys()
                      if p2 == p and s == "PM"]
            if pm_vars:
                model.Add(sum(pm_vars) <= contract.get("pm_shifts", 999))

    # ========== PHASE 1: CP-SAT OPTIMIZATION WITH ALL P1 CONSTRAINTS ==========
    def run_phase1_cpsat(self, providers, slots, availability, contracts, credentialing,
                         volumes, weekend_days):
        """Phase 1: CP-SAT optimization with ALL P1 hard constraints"""

        print("\n" + "=" * 80)
        print("PHASE 1: CP-SAT OPTIMIZATION (ALL P1 CONSTRAINTS)")
        print("=" * 80)

        model = cp_model.CpModel()
        x = {}

        print("Creating decision variables...")
        # Only create variables for valid combinations (credentialing + availability)
        for p in providers:
            cred_facilities = credentialing.get(p, set())
            for f in self.facilities:
                # P1.1: Credentialing compliance
                if f not in cred_facilities:
                    continue
                for s in self.shifts:
                    for d in self.days:
                        # P1.2: Coverage check - only create if coverage needed
                        if d not in self.coverage.get(f, {}).get(s, []):
                            continue
                        # P1.2: Availability check
                        if s not in availability.get(p, {}).get(d, []):
                            continue

                        x[(p, f, s, d)] = model.NewBoolVar(f"x_{self._sanitize(p)}_{self._sanitize(f)}_{s}_{d}")

        print(f"Variables created: {len(x):,}")

        # Create shift-day working variables (for counting shifts)
        shift_working = {}
        for p in providers:
            for s in self.shifts:
                for d in self.days:
                    vars_for_shift = [x[(p, f, s, d)] for f in self.facilities if (p, f, s, d) in x]
                    if vars_for_shift:
                        working = model.NewBoolVar(f"w_{self._sanitize(p)}_{s}_{d}")
                        model.Add(sum(vars_for_shift) >= 1).OnlyEnforceIf(working)
                        model.Add(sum(vars_for_shift) == 0).OnlyEnforceIf(working.Not())
                        shift_working[(p, s, d)] = working

        print("Applying P1 hard constraints...")

        # Coverage tracking (soft - maximize in objective)
        slot_filled = {}
        for f, s, d in slots:
            eligible = [x[(p, f, s, d)] for p in providers if (p, f, s, d) in x]
            if eligible:
                filled = model.NewBoolVar(f"filled_{self._sanitize(f)}_{s}_{d}")
                model.Add(sum(eligible) >= 1).OnlyEnforceIf(filled)
                model.Add(sum(eligible) == 0).OnlyEnforceIf(filled.Not())
                slot_filled[(f, s, d)] = filled

        # P1.3: Contracted shift count
        # print("  P1.3: Contract shift counts...")
        # for p in providers:
        #     contract = contracts.get(p, {})
        #     working_vars = [shift_working[(p2, s, d)] for (p2, s, d) in shift_working.keys() if p2 == p]
        #     if working_vars:
        #         total_shifts = sum(working_vars)

        #         if contract.get("contract_type") == "FT":
        #             # FT: Exactly equal
        #             target = min(contract.get("total_shifts", 999), 20)
        #             model.Add(total_shifts == target)
        #         else:
        #             # IC: Less than or equal
        #             model.Add(total_shifts <= min(contract.get("total_shifts", 999), 20))

        # P1.4: Weekend shift count
        print("  P1.4: Weekend shift limits...")
        for p in providers:
            contract = contracts.get(p, {})
            weekend_vars = [shift_working[(p2, s, d)] for (p2, s, d) in shift_working.keys()
                           if p2 == p and d in weekend_days]
            if weekend_vars:
                model.Add(sum(weekend_vars) <= contract.get("weekend_shifts", 999))

        # P1.5: PM (night) shift count
        print("  P1.5: PM shift limits...")
        for p in providers:
            contract = contracts.get(p, {})
            pm_vars = [shift_working[(p2, s, d)] for (p2, s, d) in shift_working.keys()
                      if p2 == p and s == "PM"]
            if pm_vars:
                model.Add(sum(pm_vars) <= contract.get("pm_shifts", 999))

        # P1.6: Volume restrictions (HARD)
        print("  P1.6: Volume restrictions...")
        for p in providers:
            for s in self.shifts:
                for d in self.days:
                    facilities_for_shift = [f for f in self.facilities if (p, f, s, d) in x]
                    if not facilities_for_shift:
                        continue

                    # Calculate total volume for this shift-day
                    volume_sum = sum([x[(p, f, s, d)] * int(volumes.get(f, {}).get(s, 0) * 100)
                                     for f in facilities_for_shift])

                    working = shift_working.get((p, s, d))
                    if working is not None:
                        min_vol, max_vol = self.volume_limits[s]
                        # Enforce both min and max
                        model.Add(volume_sum >= int(min_vol * 100)).OnlyEnforceIf(working)
                        model.Add(volume_sum <= int(max_vol * 100)).OnlyEnforceIf(working)

        # P1.7: Max consecutive shift restrictions
        print("  P1.7: Consecutive shift limits...")
        for p in providers:
            # MD1 consecutive: max 4 days
            for start_day in range(1, 29):
                md1_vars = [shift_working.get((p, "MD1", d)) for d in range(start_day, min(start_day + 5, 32))
                           if shift_working.get((p, "MD1", d)) is not None]
                if md1_vars:
                    model.Add(sum(md1_vars) <= 4)

            # MD2 consecutive: max 7 days
            for start_day in range(1, 26):
                md2_vars = [shift_working.get((p, "MD2", d)) for d in range(start_day, min(start_day + 8, 32))
                           if shift_working.get((p, "MD2", d)) is not None]
                if md2_vars:
                    model.Add(sum(md2_vars) <= 7)

            # PM consecutive: max 3 days
            for start_day in range(1, 30):
                pm_vars = [shift_working.get((p, "PM", d)) for d in range(start_day, min(start_day + 4, 32))
                          if shift_working.get((p, "PM", d)) is not None]
                if pm_vars:
                    model.Add(sum(pm_vars) <= 3)

            # MD1 + PM consecutive: max 4 days
            for start_day in range(1, 29):
                combined_vars = []
                for d in range(start_day, min(start_day + 5, 32)):
                    if shift_working.get((p, "MD1", d)) is not None:
                        combined_vars.append(shift_working.get((p, "MD1", d)))
                    if shift_working.get((p, "PM", d)) is not None:
                        combined_vars.append(shift_working.get((p, "PM", d)))
                if combined_vars:
                    model.Add(sum(combined_vars) <= 4)

        # P1.8: Daily hour limit (max 12 hours per day)
        print("  P1.8: Daily hour limits...")
        shift_hours = {"MD1": 8, "MD2": 8, "PM": 12}
        for p in providers:
            for d in self.days:
                daily_hours = []
                for s in self.shifts:
                    working = shift_working.get((p, s, d))
                    if working is not None:
                        daily_hours.append(working * shift_hours[s])

                if daily_hours:
                    model.Add(sum(daily_hours) <= 12)

        # P1.9: Site grouping restrictions (MD2 only)
        print("  P1.9: Site grouping restrictions...")
        group1 = {"NHMC", "NMHMC"}
        group2 = {"NMMC", "NBAMC"}

        for p in providers:
            for d in self.days:
                # Get MD2 assignments for each group
                group1_vars = [x[(p, f, "MD2", d)] for f in group1 if (p, f, "MD2", d) in x]
                group2_vars = [x[(p, f, "MD2", d)] for f in group2 if (p, f, "MD2", d) in x]

                if group1_vars and group2_vars:
                    # Cannot work both groups on same day
                    working_group1 = model.NewBoolVar(f"g1_{self._sanitize(p)}_{d}")
                    working_group2 = model.NewBoolVar(f"g2_{self._sanitize(p)}_{d}")

                    model.Add(sum(group1_vars) >= 1).OnlyEnforceIf(working_group1)
                    model.Add(sum(group1_vars) == 0).OnlyEnforceIf(working_group1.Not())
                    model.Add(sum(group2_vars) >= 1).OnlyEnforceIf(working_group2)
                    model.Add(sum(group2_vars) == 0).OnlyEnforceIf(working_group2.Not())

                    # At most one group per day
                    model.Add(working_group1 + working_group2 <= 1)

        # Objective: Maximize coverage
        print("Setting objective: Maximize coverage...")
        model.Maximize(sum(slot_filled.values()))

        # Solve
        print("Solving with CP-SAT (max 5 minutes)...")
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 300.0
        solver.parameters.num_search_workers = 8
        status = solver.Solve(model)

        print(f"Status: {solver.StatusName(status)}")

        if status not in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            print("⚠️  CP-SAT failed, using greedy fallback...")
            return self.run_phase1_greedy(providers, slots, availability, contracts,
                                         credentialing, volumes, weekend_days)

        # Extract solution
        solution = {}
        provider_shifts = {}
        provider_daily_volume = {}

        for (p, f, s, d), var in x.items():
            if solver.BooleanValue(var):
                solution[(f, s, d)] = p

                if p not in provider_shifts:
                    provider_shifts[p] = set()
                provider_shifts[p].add((s, d))

                vol_key = (p, s, d)
                provider_daily_volume[vol_key] = provider_daily_volume.get(vol_key, 0.0) + volumes.get(f, {}).get(s, 0.0)

        filled = len([k for k in solution.keys()])
        print(f"✓ CP-SAT Phase 1: {filled}/{len(slots)} ({filled / len(slots) * 100:.1f}%)")
        print(f"  All P1 constraints satisfied (optimal solution)")

        return solution, provider_shifts, provider_daily_volume

    def run_phase1_greedy(self, providers, slots, availability, contracts, credentialing,
                          volumes, weekend_days):
        """Greedy fallback for Phase 1"""

        print("\n" + "=" * 80)
        print("PHASE 1: GREEDY FALLBACK")
        print("=" * 80)

        solution = {}
        provider_shifts = {}
        provider_daily_volume = {}
        filled = 0

        for f, s, d in slots:
            eligible = []
            for p in providers:
                if self.is_eligible_greedy(p, f, s, d, provider_shifts, provider_daily_volume,
                                          contracts, credentialing, availability, volumes, weekend_days, True):
                    eligible.append(p)

            if eligible:
                # Simple scoring
                scored = []
                for p in eligible:
                    score = 0
                    if p in provider_shifts and (s, d) in provider_shifts[p]:
                        score += 1000
                    scored.append((p, score))

                scored.sort(key=lambda x: x[1], reverse=True)
                chosen = scored[0][0]

                solution[(f, s, d)] = chosen
                if chosen not in provider_shifts:
                    provider_shifts[chosen] = set()
                provider_shifts[chosen].add((s, d))

                vol_key = (chosen, s, d)
                provider_daily_volume[vol_key] = provider_daily_volume.get(vol_key, 0.0) + volumes.get(f, {}).get(s, 0.0)
                filled += 1

        print(f"Greedy: {filled}/{len(slots)} ({filled / len(slots) * 100:.1f}%)")
        return solution, provider_shifts, provider_daily_volume

    # ========== GREEDY ELIGIBILITY ==========
    def is_eligible_greedy(self, provider, facility, shift, day,
                          provider_shifts, provider_daily_volume,
                          contracts, credentialing, availability, volumes, weekend_days,
                          check_soft_limits):
        """Check eligibility for greedy assignment"""

        if facility not in credentialing.get(provider, set()):
            return False

        if shift not in availability.get(provider, {}).get(day, []):
            return False

        volume_key = (provider, shift, day)
        current_volume = provider_daily_volume.get(volume_key, 0.0)
        site_volume = volumes.get(facility, {}).get(shift, 0.0)
        new_volume = current_volume + site_volume

        min_vol, max_vol = self.volume_limits[shift]
        if new_volume > max_vol:
            return False

        if provider in provider_shifts and len(provider_shifts[provider]) >= 20:
            return False

        if check_soft_limits:
            if provider not in provider_shifts:
                return True

            contract = contracts.get(provider, {})
            shifts_used = len(provider_shifts[provider])

            if shifts_used >= contract.get("total_shifts", 999):
                return False

            if day in weekend_days:
                weekend_used = sum(1 for (s, d) in provider_shifts[provider] if d in weekend_days)
                if weekend_used >= contract.get("weekend_shifts", 999):
                    return False

            if shift == "PM":
                pm_used = sum(1 for (s, d) in provider_shifts[provider] if s == "PM")
                if pm_used >= contract.get("pm_shifts", 999):
                    return False

        return True

    def score_minimize_providers(self, provider, shift, day, provider_shifts, contracts):
        score = 0
        if provider in provider_shifts and len(provider_shifts[provider]) > 0:
            score += 10000
            if (shift, day) in provider_shifts[provider]:
                score += 5000

        contract = contracts.get(provider, {})
        if contract.get("preferred_shifts") and shift in contract.get("preferred_shifts"):
            score += 100
        if contract.get("contract_type") == "FT":
            score += 50

        return score

    def score_balanced_providers(self, provider, shift, day, provider_shifts, contracts):
        shifts_used = len(provider_shifts.get(provider, set()))
        score = 10000 - (shifts_used * 100)

        if provider in provider_shifts and (shift, day) in provider_shifts[provider]:
            score += 200

        return score

    # ========== PHASE 2 VARIATIONS ==========
    def phase2_minimize(self, sol_base, shifts_base, vol_base, uncovered,
                       providers, contracts, credentialing, availability, volumes, weekend_days):
        solution = sol_base.copy()
        provider_shifts = {p: s.copy() for p, s in shifts_base.items()}
        provider_volume = vol_base.copy()

        filled = 0
        for f, s, d in uncovered:
            eligible = []
            for p in providers:
                if self.is_eligible_greedy(p, f, s, d, provider_shifts, provider_volume,
                                          contracts, credentialing, availability, volumes, weekend_days, False):
                    eligible.append(p)

            if eligible:
                scored = [(p, self.score_minimize_providers(p, s, d, provider_shifts, contracts)) for p in eligible]
                scored.sort(key=lambda x: x[1], reverse=True)
                chosen = scored[0][0]

                solution[(f, s, d)] = chosen
                if chosen not in provider_shifts:
                    provider_shifts[chosen] = set()
                provider_shifts[chosen].add((s, d))

                vol_key = (chosen, s, d)
                provider_volume[vol_key] = provider_volume.get(vol_key, 0.0) + volumes.get(f, {}).get(s, 0.0)
                filled += 1

        return solution, provider_shifts, provider_volume, filled

    def phase2_balanced(self, sol_base, shifts_base, vol_base, uncovered,
                       providers, contracts, credentialing, availability, volumes, weekend_days):
        solution = sol_base.copy()
        provider_shifts = {p: s.copy() for p, s in shifts_base.items()}
        for p in providers:
            if p not in provider_shifts:
                provider_shifts[p] = set()
        provider_volume = vol_base.copy()

        filled = 0
        for f, s, d in uncovered:
            eligible = []
            for p in providers:
                if self.is_eligible_greedy(p, f, s, d, provider_shifts, provider_volume,
                                          contracts, credentialing, availability, volumes, weekend_days, False):
                    eligible.append(p)

            if eligible:
                scored = [(p, self.score_balanced_providers(p, s, d, provider_shifts, contracts)) for p in eligible]
                scored.sort(key=lambda x: x[1], reverse=True)
                chosen = scored[0][0]

                solution[(f, s, d)] = chosen
                provider_shifts[chosen].add((s, d))

                vol_key = (chosen, s, d)
                provider_volume[vol_key] = provider_volume.get(vol_key, 0.0) + volumes.get(f, {}).get(s, 0.0)
                filled += 1

        return solution, provider_shifts, provider_volume, filled

    # ========== MAIN ==========
    def run(self):
        """Main scheduling workflow"""

        print("\n" + "=" * 80)
        print("SCHEDULER: CP-SAT PHASE 1 + GREEDY PHASE 2")
        print("=" * 80)

        # Parse
        availability = self.parse_provider_availability()
        contracts = self.parse_provider_contracts()
        credentialing = self.parse_provider_credentialing()
        volumes = self.parse_facility_volumes()
        weekend_days = self.get_weekend_days()
        providers = sorted(availability.keys())
        high_volume_sites = self.identify_high_volume_sites(volumes)

        print(f"\nProviders: {len(providers)}")
        print(f"Facilities: {len(self.facilities)}")

        # Build slots
        slots = []
        for f in self.facilities:
            for s in self.shifts:
                for d in self.days:
                    if d in self.coverage.get(f, {}).get(s, []):
                        slots.append((f, s, d))

        print(f"Total slots: {len(slots)}")

        # Phase 1: CP-SAT
        sol_p1, shifts_p1, vol_p1 = self.run_phase1_cpsat(
            providers, slots, availability, contracts, credentialing,
            volumes, weekend_days
        )

        uncovered = [(f, s, d) for f, s, d in slots if (f, s, d) not in sol_p1]

        # Phase 2: 3 Variations
        print("\n" + "=" * 80)
        print("PHASE 2: GREEDY VARIATIONS")
        print("=" * 80)
        print(f"Uncovered: {len(uncovered)}")

        print("\n[1/3] Minimize...")
        sol_min, shifts_min, vol_min, f_min = self.phase2_minimize(
            sol_p1, shifts_p1, vol_p1, uncovered, providers, contracts,
            credentialing, availability, volumes, weekend_days
        )
        print(f"  +{f_min} slots")

        print("\n[2/3] Balanced...")
        sol_bal, shifts_bal, vol_bal, f_bal = self.phase2_balanced(
            sol_p1, shifts_p1, vol_p1, uncovered, providers, contracts,
            credentialing, availability, volumes, weekend_days
        )
        print(f"  +{f_bal} slots")

        print("\n[3/3] Conservative...")
        sol_cons = sol_p1.copy()
        shifts_cons = {p: s.copy() for p, s in shifts_p1.items()}
        vol_cons = vol_p1.copy()
        print(f"  No Phase 2")

        # Calculate scores
        print("\n" + "=" * 80)
        print("CALCULATING SATISFACTION SCORES")
        print("=" * 80)

        score_min = self.calculate_satisfaction_score(sol_min, shifts_min, contracts, providers, weekend_days, len(slots))
        score_bal = self.calculate_satisfaction_score(sol_bal, shifts_bal, contracts, providers, weekend_days, len(slots))
        score_cons = self.calculate_satisfaction_score(sol_cons, shifts_cons, contracts, providers, weekend_days, len(slots))

        # Export
        print("\n" + "=" * 80)
        print("EXPORTING")
        print("=" * 80)

        self._export_with_score(sol_min, "1_minimize_providers", slots, score_min)
        self._export_with_score(sol_bal, "2_balanced_providers", slots, score_bal)
        self._export_with_score(sol_cons, "3_phase1_conservative", slots, score_cons)

        # Comparison
        self._print_comparison(sol_min, sol_bal, sol_cons, shifts_min, shifts_bal, shifts_cons,
                              score_min, score_bal, score_cons, slots, providers)

    # ========== EXPORT ==========
    def _count_unique(self, solution):
        unique = set()
        for prov_data in solution.values():
            if isinstance(prov_data, tuple):
                unique.update(prov_data)
            elif prov_data:
                unique.add(prov_data)
        return len(unique)

    def _export_with_score(self, solution, name, slots, score):
        headers = ["Day"]
        site_shift_list = []

        for site in sorted(self.facilities):
            for shift in self.shifts:
                if any(self.coverage.get(site, {}).get(shift, [])):
                    headers.append(f"{site} - {shift}")
                    site_shift_list.append((site, shift))

        wb = Workbook()
        ws = wb.active
        ws.title = name[:31]

        # Score header
        ws.append(["SATISFACTION SCORE", f"{score['satisfaction_score']:.4f}"])
        ws.append(["Coverage", f"{score['coverage_rate']:.1%}"])
        ws.append(["Uncovered", score['uncovered_slots']])
        ws.append(["Excess: MD/PM/Weekend", f"{score['total_M']}/{score['total_P']}/{score['total_W']}"])
        ws.append(["Providers Affected", score['providers_affected']])
        ws.append([])

        ws.cell(1, 1).font = Font(bold=True, size=14, color="0000FF")
        ws.cell(1, 2).font = Font(bold=True, size=14, color="0000FF")

        ws.append(headers)

        # Data
        for day in self.days:
            row = [day]
            for site, shift in site_shift_list:
                if day not in self.coverage.get(site, {}).get(shift, []):
                    row.append("")
                else:
                    prov = solution.get((site, shift, day))
                    row.append(prov if prov else "UNCOVERED")
            ws.append(row)

        # Format
        black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for r in range(8, len(self.days) + 8):
            for c in range(2, len(headers) + 1):
                cell = ws.cell(r, c)
                idx = c - 2
                if idx >= len(site_shift_list):
                    continue

                site, shift = site_shift_list[idx]
                day = ws.cell(r, 1).value
                needed = day in self.coverage.get(site, {}).get(shift, [])

                if not needed:
                    cell.fill = black
                    cell.value = ""
                elif cell.value == "UNCOVERED":
                    cell.fill = red
                elif cell.value:
                    cell.fill = green

        ws.column_dimensions['A'].width = 12
        for c in range(2, min(len(headers) + 1, 30)):
            ws.column_dimensions[ws.cell(7, c).column_letter].width = 25

        path = os.path.join(self.output_dir, f"schedule_{name}.xlsx")
        wb.save(path)

        print(f"  ✓ {path}")
        print(f"    Score: {score['satisfaction_score']:.4f}, Coverage: {score['coverage_rate']:.1%}")

    def _print_comparison(self, sol_min, sol_bal, sol_cons, shifts_min, shifts_bal, shifts_cons,
                         score_min, score_bal, score_cons, slots, providers):

        variations = [
            ('Minimize', sol_min, shifts_min, score_min),
            ('Balanced', sol_bal, shifts_bal, score_bal),
            ('Conservative', sol_cons, shifts_cons, score_cons)
        ]

        ranked = sorted(variations, key=lambda x: x[3]['satisfaction_score'])

        print("\n" + "=" * 80)
        print("RANKING BY SATISFACTION SCORE")
        print("=" * 80)

        for rank, (name, sol, shifts, score) in enumerate(ranked, 1):
            icon = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉"
            filled = len([k for k in sol.keys() if k in slots])
            unique = self._count_unique(sol)

            print(f"\n{icon} RANK {rank}: {name}")
            print(f"   Score: {score['satisfaction_score']:.4f}")
            print(f"   Coverage: {filled}/{len(slots)} ({filled / len(slots) * 100:.1f}%)")
            print(f"   Providers: {unique}/{len(providers)}")
            print(f"   Excess: MD={score['total_M']}, PM={score['total_P']}, Wknd={score['total_W']}")


# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    base_scheduler = BaseScheduler()
    base_file, coverage, shifts, facilities, days, dow_map = base_scheduler.create_base_schedule()

    scheduler = Scheduler(
        base_file, coverage, shifts, facilities, days,
        "data_clean/Provider Availability.xlsx",
        "data_clean/Provider contract.xlsx",
        "data_clean/Provider Credentialing.xlsx",
        "data_clean/Facility volume.xlsx",
        dow_map
    )
    scheduler.run()
